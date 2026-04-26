"""
Per-session Excel export.

Produces one xlsx with up to 4 sheets:

    Summary       metadata + high-level analysis metrics
    TimeSeries    100 Hz unified time axis — force, per-board CoP, pose
                  (linearly interpolated), velocities, angles, angular velocities
    Pose_native   pose data at the camera's native fps (no interpolation)
    Per_rep       per-rep metrics for squat / cmj / encoder / reaction

Design:
  - Force stays at its native 100 Hz.
  - Pose is linearly interpolated to the force timeline for the TimeSeries
    sheet, but preserved un-touched in the Pose_native sheet.
  - Velocities and angular velocities use central differences on a
    6 Hz low-pass-filtered trace so numerical differentiation does not
    amplify noise.
  - Per-board CoP is computed here on demand (not persisted to forces.csv).
"""
from __future__ import annotations

import json
from pathlib import Path
from typing import Callable, Optional

import numpy as np
from scipy.signal import butter, filtfilt

import config
from src.analysis.common import ForceSession, load_force_session
from src.analysis.dispatcher import read_result
from src.analysis.pose2d import (
    ANGLE_NAMES, Pose2DSeries, get_record_start_wall_s,
    load_session_pose2d, load_video_timestamps,
)
from src.pose.mediapipe_backend import MP33_NAMES


# openpyxl is imported lazily so that importing this module (e.g. during
# app startup) does not fail when openpyxl isn't installed. The actual
# export requires it and will raise a clear error at that point.
def _require_openpyxl():
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError as e:
        raise RuntimeError(
            "openpyxl is not installed.\n"
            "Install with:\n"
            '  "C:\\Program Files\\Python311\\python.exe" -m pip install '
            '"openpyxl>=3.1"'
        ) from e
    return Workbook, Font, PatternFill, Alignment


# ─────────────────────────────────────────────────────────────────────────────
# Numerical helpers
# ─────────────────────────────────────────────────────────────────────────────

def _butter_lpf(x: np.ndarray, fs: float,
                cutoff_hz: float = 6.0, order: int = 4) -> np.ndarray:
    """Zero-phase Butterworth low-pass. Handles NaN by interpolation, then
    restores NaN at the original positions."""
    if fs <= 0 or cutoff_hz <= 0 or cutoff_hz * 2 >= fs:
        return x.copy()
    nyq = fs / 2.0
    wn = cutoff_hz / nyq
    nan_mask = np.isnan(x)
    if nan_mask.all():
        return np.full_like(x, np.nan)
    if not nan_mask.any():
        b, a = butter(order, wn, btype="low")
        return filtfilt(b, a, x)
    # Interpolate gaps so filtfilt runs, then mask back to NaN
    t_idx = np.arange(len(x), dtype=np.float64)
    valid = ~nan_mask
    x_filled = np.interp(t_idx, t_idx[valid], x[valid])
    b, a = butter(order, wn, btype="low")
    y = filtfilt(b, a, x_filled)
    y[nan_mask] = np.nan
    return y


def _velocity(x: np.ndarray, t: np.ndarray,
              filter_cutoff_hz: Optional[float] = 6.0) -> np.ndarray:
    """Derivative of x wrt t. Filters first (optional), then uses central
    differences via np.gradient. Propagates NaN from the input mask."""
    if len(x) < 2:
        return np.zeros_like(x)
    fs = 1.0 / np.mean(np.diff(t)) if len(t) > 1 else 100.0
    xf = _butter_lpf(x, fs, filter_cutoff_hz) if filter_cutoff_hz else x
    with np.errstate(invalid="ignore"):
        v = np.gradient(xf, t, edge_order=2)
    # Restore NaN where raw data was missing
    v[np.isnan(x)] = np.nan
    return v


# ─────────────────────────────────────────────────────────────────────────────
# Per-board CoP (world frame, mm) from 8 corner forces
# ─────────────────────────────────────────────────────────────────────────────

def _per_board_cop(force: ForceSession) -> dict[str, np.ndarray]:
    """Returns arrays b1_cop_x, b1_cop_y, b2_cop_x, b2_cop_y (mm)."""
    W = float(config.BOARD_WIDTH_MM)
    H = float(config.BOARD_HEIGHT_MM)
    MIN_TOTAL = 10.0       # N — matches DaqFrame._board_local_cop threshold

    def _board(corners_4: np.ndarray,
               origin_xy: tuple[float, float]) -> tuple[np.ndarray, np.ndarray]:
        # corners_4: (N, 4) = [tl, tr, bl, br] in N
        tl = np.maximum(corners_4[:, 0], 0.0)
        tr = np.maximum(corners_4[:, 1], 0.0)
        bl = np.maximum(corners_4[:, 2], 0.0)
        br = np.maximum(corners_4[:, 3], 0.0)
        tot = tl + tr + bl + br
        valid = tot >= MIN_TOTAL
        safe = np.where(tot > 0, tot, 1.0)
        cx_local = (tr * W + br * W) / safe
        cy_local = (tl * H + tr * H) / safe
        cx_world = cx_local + origin_xy[0]
        cy_world = cy_local + origin_xy[1]
        cx_world = np.where(valid, cx_world, np.nan)
        cy_world = np.where(valid, cy_world, np.nan)
        # Safety clip
        cx_world = np.clip(cx_world, origin_xy[0], origin_xy[0] + W)
        cy_world = np.clip(cy_world, origin_xy[1], origin_xy[1] + H)
        return cx_world, cy_world

    b1x, b1y = _board(force.b1, config.BOARD1_ORIGIN_MM)
    b2x, b2y = _board(force.b2, config.BOARD2_ORIGIN_MM)
    return {"b1_cop_x_mm": b1x, "b1_cop_y_mm": b1y,
            "b2_cop_x_mm": b2x, "b2_cop_y_mm": b2y}


# ─────────────────────────────────────────────────────────────────────────────
# Pose interpolation onto force timeline
# ─────────────────────────────────────────────────────────────────────────────

def _pose_to_force_grid(pose: Pose2DSeries,
                        session_dir: Path,
                        cam_id: str,
                        force_t_s: np.ndarray,
                        vis_thresh: float = 0.3) -> Optional[dict]:
    """Linearly interpolate pose x/y/visibility and pre-computed angles to
    the force sample times. Returns a dict of arrays keyed by column name,
    or None if timestamps aren't available."""
    walls = load_video_timestamps(session_dir, cam_id)
    rec_start = get_record_start_wall_s(session_dir)
    if walls is None or rec_start is None or len(walls) != len(pose):
        return None
    pose_t = (walls.astype(np.float64) - float(rec_start))
    n_force = len(force_t_s)
    n_lmk = pose.kpts_mp33.shape[1]

    out: dict[str, np.ndarray] = {}

    # Landmarks x/y
    for li in range(n_lmk):
        name = MP33_NAMES[li]
        x = pose.kpts_mp33[:, li, 0].astype(np.float64)
        y = pose.kpts_mp33[:, li, 1].astype(np.float64)
        vis = pose.vis_mp33[:, li].astype(np.float64)
        valid = (~np.isnan(x)) & (~np.isnan(y)) & (vis >= vis_thresh)
        if valid.sum() >= 2:
            out[f"{name}_x_px"] = np.interp(
                force_t_s, pose_t[valid], x[valid],
                left=np.nan, right=np.nan)
            out[f"{name}_y_px"] = np.interp(
                force_t_s, pose_t[valid], y[valid],
                left=np.nan, right=np.nan)
            out[f"{name}_vis"]  = np.interp(
                force_t_s, pose_t[valid], vis[valid],
                left=np.nan, right=np.nan)
        else:
            out[f"{name}_x_px"] = np.full(n_force, np.nan)
            out[f"{name}_y_px"] = np.full(n_force, np.nan)
            out[f"{name}_vis"]  = np.full(n_force, np.nan)

    # Angles (already in pose.angles)
    for ai, name in enumerate(pose.angle_names):
        a = pose.angles[:, ai].astype(np.float64)
        valid = ~np.isnan(a)
        if valid.sum() >= 2:
            out[f"angle_{name}_deg"] = np.interp(
                force_t_s, pose_t[valid], a[valid],
                left=np.nan, right=np.nan)
        else:
            out[f"angle_{name}_deg"] = np.full(n_force, np.nan)

    return out


# ─────────────────────────────────────────────────────────────────────────────
# Styling helpers (built lazily, see _styles())
# ─────────────────────────────────────────────────────────────────────────────

def _styles():
    _, Font, PatternFill, Alignment = _require_openpyxl()
    return {
        "header_font": Font(bold=True, color="FFFFFF"),
        "header_fill": PatternFill(start_color="2E7D32",
                                    end_color="2E7D32",
                                    fill_type="solid"),
        "key_font":    Font(bold=True),
        "center":      Alignment(horizontal="center"),
    }


def _style_header(row, styles) -> None:
    for cell in row:
        cell.font = styles["header_font"]
        cell.fill = styles["header_fill"]
        cell.alignment = styles["center"]


def _autosize(ws, min_w: int = 10, max_w: int = 22) -> None:
    for col_cells in ws.columns:
        first = col_cells[0]
        col_letter = first.column_letter
        longest = max(
            (len(str(c.value)) if c.value is not None else 0)
            for c in col_cells[:50]   # sample first 50 rows for speed
        )
        ws.column_dimensions[col_letter].width = max(min_w,
                                                     min(max_w, longest + 2))


# ─────────────────────────────────────────────────────────────────────────────
# Sheet writers
# ─────────────────────────────────────────────────────────────────────────────

def _write_summary(wb, session_dir: Path,
                   force: ForceSession,
                   pose: Optional[Pose2DSeries],
                   styles: dict) -> None:
    ws = wb.create_sheet("Summary")
    meta = {}
    try:
        meta = json.loads(
            (session_dir / "session.json").read_text(encoding="utf-8"))
    except Exception:
        pass
    result = read_result(session_dir) or {}
    res_body = result.get("result") or {}

    ws.append(["Key", "Value"])
    _style_header(ws[1], styles)

    def row(k, v): ws.append([k, v])

    row("session_name",        session_dir.name)
    row("test",                meta.get("test"))
    row("stance",              meta.get("stance"))
    row("vision",              meta.get("vision"))
    row("duration_s",          meta.get("duration_s"))
    row("cancelled",           meta.get("cancelled"))
    row("fell_off_detected",   meta.get("fell_off_detected"))
    row("wait_duration_s",     meta.get("wait_duration_s"))
    row("record_start_wall_s", meta.get("record_start_wall_s"))
    row("subject_id",          meta.get("subject_id"))
    row("subject_name",        meta.get("subject_name"))
    row("subject_kg",          meta.get("subject_kg"))
    row("n_daq_samples",       int(meta.get("n_daq_samples") or 0))
    row("n_stimuli",           int(meta.get("n_stimuli") or 0))

    ws.append([])
    ws.append(["Analysis result (key metrics)", ""])
    _style_header(ws[ws.max_row], styles)
    # Flatten the top-level numeric / string result keys
    if result.get("error"):
        row("analysis_error", result.get("error"))
    for k, v in res_body.items():
        if isinstance(v, (int, float, str)) or v is None:
            row(k, v)

    # Camera / pose meta
    ws.append([])
    ws.append(["Pose / camera", ""])
    _style_header(ws[ws.max_row], styles)
    if pose is not None:
        row("pose_cam_id",     pose.cam_id)
        row("pose_frames",     len(pose))
        row("pose_fps",        pose.fps)
        row("pose_backend",    pose.backend)
        row("pose_complexity", pose.model_complexity)
        row("pose_image_w",    pose.image_size[0])
        row("pose_image_h",    pose.image_size[1])
    else:
        row("pose_cam_id",     "(no pose data)")

    row("camera_rotation_deg", getattr(config, "CAMERA_ROTATION", 0))
    row("camera_mirror",       bool(getattr(config, "CAMERA_MIRROR", False)))

    # Force row-bolding
    for r in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=1):
        for c in r:
            if c.value:
                c.font = styles["key_font"]
    _autosize(ws)


def _write_timeseries(wb, session_dir: Path,
                      force: ForceSession,
                      pose: Optional[Pose2DSeries],
                      cam_id: Optional[str],
                      styles: dict,
                      progress_cb: Optional[Callable[[str], None]] = None
                      ) -> None:
    ws = wb.create_sheet("TimeSeries")
    t = force.t_s.astype(np.float64)
    n = len(t)

    # 1) Build columns
    cols: dict[str, np.ndarray] = {}
    cols["t_s"]          = t
    cols["t_wall_s"]     = np.arange(n, dtype=np.float64)   # placeholder
    # Recover actual t_wall from forces.csv if available
    try:
        import pandas as pd
        df = pd.read_csv(session_dir / "forces.csv", usecols=["t_wall"])
        if len(df) == n:
            cols["t_wall_s"] = df["t_wall"].to_numpy(dtype=np.float64)
    except Exception:
        pass

    # Force raw (8 corners)
    corner_names = ["b1_tl_N", "b1_tr_N", "b1_bl_N", "b1_br_N",
                    "b2_tl_N", "b2_tr_N", "b2_bl_N", "b2_br_N"]
    for i, name in enumerate(corner_names[:4]):
        cols[name] = force.b1[:, i].astype(np.float64)
    for i, name in enumerate(corner_names[4:]):
        cols[name] = force.b2[:, i].astype(np.float64)

    cols["b1_total_N"] = force.b1_total.astype(np.float64)
    cols["b2_total_N"] = force.b2_total.astype(np.float64)
    cols["total_N"]    = force.total.astype(np.float64)

    # Encoder
    cols["enc1_mm"] = force.enc1.astype(np.float64)
    cols["enc2_mm"] = force.enc2.astype(np.float64)

    # CoP per-board + combined
    pb = _per_board_cop(force)
    for k, v in pb.items():
        cols[k] = v
    cols["cop_x_mm"] = force.cop_x.astype(np.float64)
    cols["cop_y_mm"] = force.cop_y.astype(np.float64)

    # Pose interpolated to force grid
    pose_cols: dict[str, np.ndarray] = {}
    if pose is not None and cam_id is not None:
        if progress_cb:
            progress_cb("pose: interpolating to force timeline")
        grid = _pose_to_force_grid(pose, session_dir, cam_id, t)
        if grid is not None:
            pose_cols = grid

    # Velocities (landmark px/s) + angular velocities (deg/s)
    if pose_cols:
        if progress_cb:
            progress_cb("pose: computing velocities")
        for name in MP33_NAMES:
            x = pose_cols.get(f"{name}_x_px")
            y = pose_cols.get(f"{name}_y_px")
            if x is None or y is None:
                continue
            pose_cols[f"{name}_vx_px_s"] = _velocity(x, t, filter_cutoff_hz=6.0)
            pose_cols[f"{name}_vy_px_s"] = _velocity(y, t, filter_cutoff_hz=6.0)
        for name in ANGLE_NAMES:
            a = pose_cols.get(f"angle_{name}_deg")
            if a is None:
                continue
            pose_cols[f"angle_vel_{name}_deg_s"] = _velocity(
                a, t, filter_cutoff_hz=6.0)

    # Insert pose cols with grouped ordering for readability:
    #   1) coords (x, y, vis) for each landmark
    #   2) velocities (vx, vy) for each landmark
    #   3) angles
    #   4) angular velocities
    def _add(group_keys):
        for k in group_keys:
            if k in pose_cols:
                cols[k] = pose_cols[k]

    for name in MP33_NAMES:
        _add([f"{name}_x_px", f"{name}_y_px", f"{name}_vis"])
    for name in MP33_NAMES:
        _add([f"{name}_vx_px_s", f"{name}_vy_px_s"])
    for name in ANGLE_NAMES:
        _add([f"angle_{name}_deg"])
    for name in ANGLE_NAMES:
        _add([f"angle_vel_{name}_deg_s"])

    # 2) Write header
    headers = list(cols.keys())
    ws.append(headers)
    _style_header(ws[1], styles)
    ws.freeze_panes = "C2"

    # 3) Write data rows
    if progress_cb:
        progress_cb(f"writing {n} rows × {len(headers)} cols")
    # Pre-stack into a 2D list once — avoids repeated dict access per cell
    matrix = np.column_stack([cols[h] for h in headers])
    for i in range(n):
        row = matrix[i].tolist()
        # Convert NaN -> None for Excel (empty cell)
        row = [None if (isinstance(v, float) and np.isnan(v)) else v
               for v in row]
        ws.append(row)
    # Fix column widths (just sample — 230 columns otherwise too slow)
    for cell in ws[1][:15]:
        ws.column_dimensions[cell.column_letter].width = 14


def _write_pose_native(wb, pose: Pose2DSeries,
                      session_dir: Path, cam_id: Optional[str],
                      styles: dict) -> None:
    ws = wb.create_sheet("Pose_native")
    # Reconstruct force-time for each pose frame (via wall timestamps)
    walls = load_video_timestamps(session_dir, cam_id) if cam_id else None
    rec_start = get_record_start_wall_s(session_dir)
    if walls is not None and rec_start is not None and len(walls) == len(pose):
        t_s = (walls.astype(np.float64) - float(rec_start))
    else:
        t_s = np.arange(len(pose), dtype=np.float64) / max(pose.fps, 1.0)

    headers = ["frame_idx", "t_s"]
    for n in MP33_NAMES:
        headers += [f"{n}_x_px", f"{n}_y_px", f"{n}_vis"]
    for n in pose.angle_names:
        headers += [f"angle_{n}_deg"]

    ws.append(headers)
    _style_header(ws[1], styles)
    ws.freeze_panes = "C2"

    n_frames = len(pose)
    for i in range(n_frames):
        row = [i, float(t_s[i])]
        for li in range(33):
            x = float(pose.kpts_mp33[i, li, 0])
            y = float(pose.kpts_mp33[i, li, 1])
            v = float(pose.vis_mp33[i, li])
            row += [None if np.isnan(x) else x,
                    None if np.isnan(y) else y,
                    None if np.isnan(v) else v]
        for ai in range(pose.angles.shape[1]):
            a = float(pose.angles[i, ai])
            row.append(None if np.isnan(a) else a)
        ws.append(row)
    for cell in ws[1][:15]:
        ws.column_dimensions[cell.column_letter].width = 14


def _write_per_rep(wb, session_dir: Path, styles: dict) -> None:
    """Optional per-rep table — only useful for squat/cmj/encoder."""
    result = read_result(session_dir) or {}
    body = result.get("result") or {}
    reps = body.get("reps")
    if not reps or not isinstance(reps, list):
        return

    ws = wb.create_sheet("Per_rep")
    # Collect all scalar keys across reps
    keys = []
    for r in reps:
        for k, v in r.items():
            if isinstance(v, (int, float, str)) or v is None:
                if k not in keys:
                    keys.append(k)
    headers = ["rep_idx"] + keys
    ws.append(headers)
    _style_header(ws[1], styles)
    for i, r in enumerate(reps):
        row = [i + 1]
        for k in keys:
            v = r.get(k)
            if isinstance(v, float) and np.isnan(v):
                v = None
            row.append(v)
        ws.append(row)
    _autosize(ws)


# ─────────────────────────────────────────────────────────────────────────────
# Entry point
# ─────────────────────────────────────────────────────────────────────────────

def export_session_xlsx(session_dir: str | Path,
                        out_path: Optional[str | Path] = None,
                        progress_cb: Optional[Callable[[str], None]] = None
                        ) -> Path:
    """Export one session to an xlsx. Returns the saved path."""
    session_dir = Path(session_dir)
    if not session_dir.exists():
        raise FileNotFoundError(session_dir)
    if out_path is None:
        out_path = session_dir / f"{session_dir.name}.xlsx"
    else:
        out_path = Path(out_path)

    # Lazy-load openpyxl only at the actual export call (not at module
    # import time). This keeps the app startable even when openpyxl is
    # missing; the error surfaces only if the user clicks "Excel export".
    Workbook, _Font, _PatternFill, _Alignment = _require_openpyxl()
    styles = _styles()

    if progress_cb: progress_cb("loading force data")
    force = load_force_session(session_dir)
    pose_map = load_session_pose2d(session_dir)
    cam_id = next(iter(pose_map.keys())) if pose_map else None
    pose = pose_map.get(cam_id) if cam_id else None
    if progress_cb:
        progress_cb(f"force={len(force.t_s)} samples "
                    f"{'pose=' + str(len(pose)) + ' frames' if pose else 'no pose'}")

    wb = Workbook()
    _write_summary(wb, session_dir, force, pose, styles)
    _write_timeseries(wb, session_dir, force, pose, cam_id, styles, progress_cb)
    if pose is not None:
        if progress_cb: progress_cb("writing Pose_native sheet")
        _write_pose_native(wb, pose, session_dir, cam_id, styles)
    if progress_cb: progress_cb("writing Per_rep sheet")
    _write_per_rep(wb, session_dir, styles)
    # Remove the default empty sheet
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    if progress_cb: progress_cb(f"saving to {out_path}")
    wb.save(str(out_path))
    return out_path
